#!/usr/bin/env python3
"""Validate every generated sync output before Git is allowed to publish it."""

from __future__ import annotations

import json
import hashlib
import math
import re
import sys
import zipfile
from itertools import zip_longest
from pathlib import Path

import openpyxl


HERE = Path(__file__).resolve().parent
PROJECT_ROOT = HERE.parent
sys.path.insert(0, str(HERE))
sys.path.insert(0, str(HERE / "phase2"))

from regen_website_data import (  # noqa: E402
    _find_code_desc_columns,
    _looks_like_part_code,
    build_metals_products,
    build_mini_products,
    read_partsbook_prices,
)
from build_lookup import build_lookup_rows  # noqa: E402
from metal_matching import MetalMatcher  # noqa: E402
from wire_catalogue import build_indexes, is_data_sheet  # noqa: E402


MINI_TS = PROJECT_ROOT / "lib" / "mini-data.ts"
METALS_TS = PROJECT_ROOT / "lib" / "metals-data.ts"
MINI_BOOK = PROJECT_ROOT / "final-deliverables" / "Mini Catalogue Self Updating.xlsm"
METALS_BOOK = PROJECT_ROOT / "final-deliverables" / "Metals catalogue 2023.xlsx"
MINI_INVOICE = PROJECT_ROOT / "final-deliverables" / "Mini Invoice Template.xlsm"
METALS_INVOICE = PROJECT_ROOT / "final-deliverables" / "Metals Invoice.xlsm"
MINI_INVOICE_SOURCE = PROJECT_ROOT / "data-source" / "Mini Invoice Template.xlsm"
METALS_INVOICE_SOURCE = PROJECT_ROOT / "data-source" / "Metals Invoice.xlsm"
PARTSBOOK = PROJECT_ROOT / "data-source" / "PartsbookBenji2014.xlsx"
MINI_PDF = PROJECT_ROOT / "public" / "catalogue" / "mini-catalogue.pdf"
METALS_PDF = PROJECT_ROOT / "public" / "catalogue" / "metals-catalogue.pdf"
CATALOGUE_VERSIONS = PROJECT_ROOT / "lib" / "catalogue-versions.ts"

INVOICE_PRICE_SHEETS = {
    "Parts Data": "_PriceLookup",
    "KDMSPC": "_KDMSPC",
    "MSPORT": "_MSPORT",
    "Magnum": "_Magnum",
    "Somerford": "_Somerford",
}


def close_enough(left, right) -> bool:
    if left is None or right is None:
        return left is right
    if isinstance(left, (int, float)) and isinstance(right, (int, float)):
        return math.isclose(float(left), float(right), rel_tol=0, abs_tol=0.001)
    return str(left).strip() == str(right).strip()


def read_exported_array(path: Path, marker: str):
    source = path.read_text(encoding="utf-8")
    start = source.find(marker)
    if start < 0:
        raise ValueError(f"{path.name}: missing marker {marker!r}")
    start = source.find("=", start + len(marker))
    start = source.find("[", start + 1)
    if start < 0:
        raise ValueError(f"{path.name}: missing generated JSON array")
    return json.JSONDecoder().raw_decode(source[start:])[0]


def validate_website_data(failures: list[str]) -> None:
    expected_mini = build_mini_products()
    actual_mini = read_exported_array(MINI_TS, "export const products")
    if len(actual_mini) != len(expected_mini):
        failures.append(
            f"website mini count is {len(actual_mini)}, expected {len(expected_mini)}"
        )
    for index, (expected, actual) in enumerate(zip(expected_mini, actual_mini), 1):
        for field in ("code", "name", "section", "priceExVat", "priceIncVat"):
            if not close_enough(expected.get(field), actual.get(field)):
                failures.append(
                    f"website mini row {index} {field}: "
                    f"{actual.get(field)!r} != {expected.get(field)!r}"
                )
                break

    expected_metals = build_metals_products()
    actual_metals = read_exported_array(METALS_TS, "export const metals")
    if len(actual_metals) != len(expected_metals):
        failures.append(
            f"website metals count is {len(actual_metals)}, expected {len(expected_metals)}"
        )
    for index, (expected, actual) in enumerate(zip(expected_metals, actual_metals), 1):
        for field in (
            "sourceSheet",
            "metal",
            "form",
            "spec",
            "size",
            "unit",
            "priceExVat",
            "priceIncVat",
        ):
            if not close_enough(expected.get(field), actual.get(field)):
                failures.append(
                    f"website metals row {index} {field}: "
                    f"{actual.get(field)!r} != {expected.get(field)!r}"
                )
                break


def validate_mini_workbook(failures: list[str]) -> int:
    prices = read_partsbook_prices()

    with zipfile.ZipFile(MINI_BOOK) as archive:
        names = set(archive.namelist())
        if any(name.startswith("xl/externalLinks/") for name in names):
            failures.append("mini workbook still contains external-link files")
        for name in names:
            if not re.fullmatch(r"xl/worksheets/sheet\d+\.xml", name):
                continue
            if b"[1]Parts Data" in archive.read(name):
                failures.append(
                    f"mini workbook still contains a Partsbook formula in {name}"
                )
                break

    wb = openpyxl.load_workbook(MINI_BOOK, data_only=True, keep_vba=True)
    if "_PriceLookup" in wb.sheetnames:
        failures.append("mini workbook unexpectedly contains _PriceLookup")

    checked = 0
    for sheet_name in wb.sheetnames:
        if not (
            (sheet_name.endswith("B") and sheet_name[:-1].isdigit())
            or sheet_name in {"APX1", "APX2"}
        ):
            continue
        ws = wb[sheet_name]
        header = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
        for code_col, _desc_col in _find_code_desc_columns(header):
            price_col = code_col + 2
            for row_idx in range(2, ws.max_row + 1):
                raw_code = ws.cell(row_idx, code_col).value
                code = str(raw_code).strip() if raw_code is not None else ""
                if not _looks_like_part_code(code):
                    continue
                expected = prices.get(code)
                if expected is None:
                    continue
                actual = ws.cell(row_idx, price_col).value
                checked += 1
                if not close_enough(actual, expected):
                    failures.append(
                        f"mini workbook {sheet_name}!{ws.cell(row_idx, price_col).coordinate} "
                        f"for {code}: {actual!r} != {expected!r}"
                    )
    wb.close()
    return checked


def validate_metals_workbook(failures: list[str]) -> int:
    rows, _collisions = build_lookup_rows()
    keys = build_indexes(rows)
    matcher = MetalMatcher(rows)
    missing_targets = sorted(
        link_id
        for link_id in matcher.links.values()
        if link_id not in matcher.by_link_id
    )
    if missing_targets:
        failures.append(
            f"{len(missing_targets)} saved metal links point to removed master rows"
        )

    wb = openpyxl.load_workbook(METALS_BOOK, data_only=True)
    linked = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if not is_data_sheet(ws):
            continue
        for row_idx in range(2, ws.max_row + 1):
            lookup_key = ws.cell(row_idx, 11).value
            if not lookup_key:
                continue
            linked += 1
            master = keys.get(str(lookup_key))
            if not master:
                failures.append(
                    f"metals workbook {sheet_name}!K{row_idx} has an unknown lookup key"
                )
                continue
            actual = ws.cell(row_idx, 5).value
            expected = master["priceEx"]
            if not close_enough(actual, expected):
                failures.append(
                    f"metals workbook {sheet_name}!E{row_idx}: "
                    f"{actual!r} != {expected!r} from {master['src']}"
                )
    wb.close()

    if linked != len(matcher.links):
        failures.append(
            f"metals workbook has {linked} linked rows, saved map has {len(matcher.links)}"
        )
    return linked


def validate_invoices(failures: list[str]) -> int:
    if not MINI_INVOICE.exists():
        failures.append("Mini Invoice Template.xlsm was not generated")
        return 0
    if not METALS_INVOICE.exists():
        failures.append("Metals Invoice.xlsm was not copied")
    elif METALS_INVOICE_SOURCE.exists():
        if hashlib.sha256(METALS_INVOICE.read_bytes()).digest() != hashlib.sha256(
            METALS_INVOICE_SOURCE.read_bytes()
        ).digest():
            failures.append("Metals Invoice.xlsm is not an unchanged source copy")

    checked = 0
    partsbook = openpyxl.load_workbook(PARTSBOOK, data_only=True, read_only=True)
    invoice = openpyxl.load_workbook(
        MINI_INVOICE, data_only=True, keep_vba=True, read_only=True
    )
    for source_name, embedded_name in INVOICE_PRICE_SHEETS.items():
        if embedded_name not in invoice.sheetnames:
            failures.append(f"Mini invoice is missing hidden sheet {embedded_name}")
            continue
        embedded = invoice[embedded_name]
        if embedded.sheet_state != "hidden":
            failures.append(f"Mini invoice sheet {embedded_name} is not hidden")
        source = partsbook[source_name]
        source_rows = source.iter_rows(values_only=True)
        embedded_rows = embedded.iter_rows(values_only=True)
        for row_idx, (source_row, embedded_row) in enumerate(
            zip_longest(source_rows, embedded_rows, fillvalue=()), 1
        ):
            for col_idx, expected in enumerate(source_row, 1):
                if expected is None or expected == "":
                    continue
                checked += 1
                actual = (
                    embedded_row[col_idx - 1]
                    if col_idx <= len(embedded_row)
                    else None
                )
                if not close_enough(actual, expected):
                    failures.append(
                        f"Mini invoice {embedded_name} row {row_idx}, col {col_idx}: "
                        f"{actual!r} != {expected!r} from Partsbook {source_name}"
                    )
                    if len(failures) >= 40:
                        break
            if len(failures) >= 40:
                break
    invoice.close()
    partsbook.close()

    with zipfile.ZipFile(MINI_INVOICE) as archive:
        names = archive.namelist()
        if "xl/vbaProject.bin" not in names:
            failures.append("Mini invoice lost its VBA project")
        external_files = [
            name for name in names if name.startswith("xl/externalLinks/")
        ]
        if external_files:
            failures.append("Mini invoice still contains external workbook links")
        external_formulas = 0
        internal_formulas = 0
        for name in names:
            if not (
                name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
            ):
                continue
            source = archive.read(name).decode("utf-8")
            external_formulas += source.count("[1]")
            internal_formulas += sum(
                source.count(sheet_name)
                for sheet_name in INVOICE_PRICE_SHEETS.values()
            )
        if external_formulas:
            failures.append(
                f"Mini invoice still has {external_formulas} external formula references"
            )
        if internal_formulas == 0:
            failures.append("Mini invoice has no formulas using embedded prices")

    return checked


def validate_pdfs(failures: list[str]) -> None:
    for path in (MINI_PDF, METALS_PDF):
        if not path.exists() or path.stat().st_size < 10_000:
            failures.append(f"{path.name} was not generated correctly")
            continue
        if path.stat().st_mtime < min(MINI_BOOK.stat().st_mtime, METALS_BOOK.stat().st_mtime) - 5:
            failures.append(f"{path.name} is older than the generated customer workbooks")


def validate_catalogue_versions(failures: list[str]) -> None:
    if not CATALOGUE_VERSIONS.exists():
        failures.append("catalogue-versions.ts was not generated")
        return

    source = CATALOGUE_VERSIONS.read_text(encoding="utf-8")
    expected = {
        "miniCatalogueVersion": hashlib.sha256(MINI_PDF.read_bytes()).hexdigest()[:16],
        "metalsCatalogueVersion": hashlib.sha256(METALS_PDF.read_bytes()).hexdigest()[:16],
    }
    for name, digest in expected.items():
        match = re.search(rf'export const {name} = "([a-f0-9]+)";', source)
        if not match:
            failures.append(f"catalogue-versions.ts is missing {name}")
        elif match.group(1) != digest:
            failures.append(
                f"{name} is {match.group(1)}, expected current PDF hash {digest}"
            )


def main() -> None:
    failures: list[str] = []
    print("Validating generated website data and customer files")
    validate_website_data(failures)
    mini_checked = validate_mini_workbook(failures)
    metals_checked = validate_metals_workbook(failures)
    invoice_checked = validate_invoices(failures)
    validate_pdfs(failures)
    validate_catalogue_versions(failures)

    if failures:
        print()
        print(f"VALIDATION FAILED ({len(failures)} problems)")
        for failure in failures[:40]:
            print(f"  - {failure}")
        if len(failures) > 40:
            print(f"  - ... plus {len(failures) - 40} more")
        sys.exit(1)

    print(
        f"  OK {mini_checked} calculated Mini prices, "
        f"{metals_checked} linked metals prices, {invoice_checked} embedded "
        "Mini invoice cells, website data, customer workbooks, and both PDFs"
    )


if __name__ == "__main__":
    main()
