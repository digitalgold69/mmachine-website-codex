#!/usr/bin/env python3
"""
Regenerate lib/mini-data.ts and lib/metals-data.ts from the customer-facing
catalogue files (Mini Catalogue Self Updating + Metals catalogue 2023) joined
with master prices (PartsbookBenji + Metals.xlsx).

This is the v2 of sync-data — flipped from the original. The customer-facing
docs are now the "what to display" definition. Master files are price-only.

Workflow:
  1. Read Mini Catalogue Self Updating → list of part codes by section
     (sections are encoded in sheet names: 120B, 130B, ..., 510B + APX1, APX2)
  2. Look up each code's price from PartsbookBenji's Parts Data.
  3. For metals: read Metals catalogue 2023 → list of (shape, metal, spec, size)
     rows. Look up each row's price from Metals.xlsx via the same fuzzy
     matching used by wire_catalogue.py.
  4. Write lib/mini-data.ts and lib/metals-data.ts.

Result: the website shows exactly what's in the customer-facing docs, with
prices that auto-update each time the masters change.
"""
import json
import openpyxl
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE / "phase2"))

PROJECT_ROOT = HERE.parent
SRC = PROJECT_ROOT / "data-source"
LIB = PROJECT_ROOT / "lib"
PUBLIC = PROJECT_ROOT / "public"

FEATURED_JSON = SRC / "featured-work.json"
FEATURED_IMG_SRC = SRC / "featured-images"
FEATURED_IMG_DST = PUBLIC / "featured"

# Mini Catalogue is required for product list; PartsbookBenji for prices
PARTSBOOK    = SRC / "PartsbookBenji2014.xlsx"
MINI_CAT     = SRC / "Mini Catalogue Self Updating.xlsm"

# Metals catalogue is required for product list; Metals.xlsx for prices
METALS_CAT   = SRC / "Metals catalogue 2023.xlsx"
METALS       = SRC / "Metals.xlsx"

from build_lookup import build_lookup_rows
from wire_catalogue import build_indexes, find_match, HIGH_CONFIDENCE, s as strip_str, is_data_sheet


# ─────────────────────────────────────────────────────────────────────────────
# SECTIONS — same list as the original sync-data.mjs (from the printed PDF)
# ─────────────────────────────────────────────────────────────────────────────

SECTIONS = [
    {"code": "120", "label": "REAR PANEL",         "subtitle": "Rear doors, Tailgate, Cab & Bulkhead",                 "order":  1, "mode": "exterior"},
    {"code": "130", "label": "SIDE REPAIRS",       "subtitle": "Apex Panels, Lamp Panels, Door Steps",                 "order":  2, "mode": "exterior"},
    {"code": "140", "label": "SIDE REPAIRS",       "subtitle": "Traveller Wood Kit, Quarter Panels",                   "order":  3, "mode": "exterior"},
    {"code": "150", "label": "DOORS",              "subtitle": "Door Skins Repairs, Window Rails",                     "order":  4, "mode": "exterior"},
    {"code": "160", "label": "FRONT & ROOF",       "subtitle": "Bonnet, Outer Wings, Roof, Scuttle Panel",             "order":  5, "mode": "exterior"},
    {"code": "170", "label": "FRONT PANEL",        "subtitle": "Number Plate Hanger, Stiffener, Adaptors",             "order":  6, "mode": "exterior"},
    {"code": "210", "label": "FRONT INNER WINGS",  "subtitle": "Inner Wings, Vent Panel, Apex Panels",                 "order":  7, "mode": "interior"},
    {"code": "220", "label": "FRONT BULKHEAD",     "subtitle": "Heating & Cooling, Radiator Cowls",                    "order":  8, "mode": "interior"},
    {"code": "230", "label": "FRONT BULKHEAD",     "subtitle": "Crossmember, Repair Panels, Brackets",                 "order":  9, "mode": "interior"},
    {"code": "310", "label": "FLOOR PANELS",       "subtitle": "Full Floor Sections and Sills",                        "order": 10, "mode": "interior"},
    {"code": "320", "label": "FLOOR PANELS",       "subtitle": "Floor Repair Panels, Half Floors, Tunnels",            "order": 11, "mode": "interior"},
    {"code": "330", "label": "FLOOR PANELS",       "subtitle": "Flat Floors, Van, Traveller, Pick-ups",                "order": 12, "mode": "interior"},
    {"code": "340", "label": "FLOOR ASSEMBLY",     "subtitle": "Crossmember, Companion Box, Mounts",                   "order": 13, "mode": "interior"},
    {"code": "350", "label": "INNER FLOOR",        "subtitle": "Inner Sills, Heel Board, Tunnel Cover",                "order": 14, "mode": "interior"},
    {"code": "410", "label": "REAR FLOOR",         "subtitle": "Boot Floor, Rear Floor, Rear Quarters",                "order": 15, "mode": "interior"},
    {"code": "420", "label": "REAR FLOOR",         "subtitle": "Boot Floor Pieces, Spare Wheel Wells",                 "order": 16, "mode": "interior"},
    {"code": "430", "label": "REAR FLOOR",         "subtitle": "Rear Repair Panels, Estate, Pick-ups",                 "order": 17, "mode": "interior"},
    {"code": "510", "label": "SUBFRAMES",          "subtitle": "Front & Rear Subframes, Servo Brackets",               "order": 18, "mode": "interior"},
    {"code": "Apx1","label": "ACCESSORIES",        "subtitle": "Door Seals, Dash Vinyl, Grommets Etc",                 "order": 19, "mode": "interior"},
    {"code": "Apx2","label": "SWITCH PANELS",      "subtitle": "Switch Panel Variations",                              "order": 20, "mode": "interior"},
]

METALS_CATEGORIES = [
    {"key": "tool_steel",  "label": "Tool steels"},
    {"key": "mild_steel",  "label": "Mild / carbon steel"},
    {"key": "stainless",   "label": "Stainless steel"},
    {"key": "aluminium",   "label": "Aluminium"},
    {"key": "brass",       "label": "Brass"},
    {"key": "bronze",      "label": "Bronze"},
    {"key": "copper",      "label": "Copper"},
    {"key": "cast_iron",   "label": "Cast iron"},
    {"key": "plastics",    "label": "Plastics"},
    {"key": "misc",        "label": "Misc"},
]


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def parse_description(desc):
    """Best-effort parse of fitment metadata from the description string."""
    d = str(desc) if desc else ""
    hand = "unhanded"
    if re.search(r"\bLH\b|\bL\.H\.?\b", d): hand = "LH assembly"
    elif re.search(r"\bRH\b|\bR\.H\.?\b", d): hand = "RH assembly"

    mark = None
    m = re.search(r"\bMk\s*([0-9IVX]+(?:\s*[-–]\s*[0-9IVX]+)?(?:\s*on)?)", d, re.I)
    if m:
        raw = m.group(1).replace(" ", "").replace("–", "-")
        mark = "Mark " + raw.upper()

    bodyType = None
    if re.search(r"\bSaloon\b", d, re.I): bodyType = "Saloon"
    elif re.search(r"\bTraveller\b|\bTrav\b", d, re.I): bodyType = "Traveller"
    elif re.search(r"\bVan\b", d, re.I): bodyType = "Van"
    elif re.search(r"\bPick[- ]?Up\b|\bPick-?up\b", d, re.I): bodyType = "Pick-Up"
    elif re.search(r"\bEstate\b", d, re.I): bodyType = "Estate"
    elif re.search(r"\bClubman\b", d, re.I): bodyType = "Clubman"
    elif re.search(r"\bCooper\b", d, re.I): bodyType = "Cooper"
    return bodyType, mark, hand


def fits_string(bodyType, mark):
    parts = [p for p in [bodyType, mark] if p]
    return ", ".join(parts) if parts else "All"


def round2(x):
    return round(x * 100) / 100 if x is not None else None


def category_from_metal_sheet(name):
    n = name.lower()
    if n.startswith("alu"): return "aluminium"
    if n.startswith("brass"): return "brass"
    if "bronze" in n or n.startswith("ph brnz") or n.startswith("pb"): return "bronze"
    if n.startswith("cu") or "copper" in n or "nickel silver" in n or "nil ag" in n: return "copper"
    if "stainless" in n or n.startswith("st steel") or n.startswith(" stainless") or n.startswith("st st"): return "stainless"
    if n.startswith("steel") or "gauge plate" in n or "silver steel" in n: return "mild_steel"
    if "cast" in n: return "cast_iron"
    if "plastic" in n: return "plastics"
    if "lg" in n or "colphos" in n: return "bronze"
    return "misc"


# ─────────────────────────────────────────────────────────────────────────────
# Mini products — read product list from Mini Catalogue, prices from PartsbookBenji
# ─────────────────────────────────────────────────────────────────────────────

def read_mini_catalogue_codes():
    """Walk Mini Catalogue's B-sheets (120B…510B) and APX1/APX2.
    Returns a list of (code, description, section) tuples preserving the
    catalogue's left-then-right reading order."""
    if not MINI_CAT.exists():
        print(f"  ! Mini Catalogue not found at {MINI_CAT}; mini products will be empty")
        return []

    wb = openpyxl.load_workbook(MINI_CAT, data_only=True)
    out = []

    # Section sheets (120B, 130B, ...): left list cols B/C, right list cols H/I
    for sn in wb.sheetnames:
        m = re.fullmatch(r"(\d+)B", sn)
        if not m: continue
        section = m.group(1)
        ws = wb[sn]
        for r in range(2, ws.max_row + 1):
            # left
            code = strip_str(ws.cell(r, 2).value)
            desc = strip_str(ws.cell(r, 3).value)
            if code and desc:
                out.append((code, desc, section))
            # right
            code = strip_str(ws.cell(r, 8).value)
            desc = strip_str(ws.cell(r, 9).value)
            if code and desc:
                out.append((code, desc, section))

    # APX sheets — same layout but different section labels
    for sn, section in [("APX1", "Apx1"), ("APX2", "Apx2")]:
        if sn not in wb.sheetnames: continue
        ws = wb[sn]
        for r in range(2, ws.max_row + 1):
            for code_col, desc_col in [(2, 3), (8, 9), (1, 2), (5, 6)]:
                code = strip_str(ws.cell(r, code_col).value)
                desc = strip_str(ws.cell(r, desc_col).value)
                if code and desc and re.match(r"^[\d.]+(?:\.\d+)?$|^[A-Z0-9-]{3,}$", code):
                    out.append((code, desc, section))
                    break  # only one product per row pair

    return out


def read_partsbook_prices():
    """Read PartsbookBenji's Parts Data → dict {code → price}."""
    wb = openpyxl.load_workbook(PARTSBOOK, data_only=True)
    ws = wb["Parts Data"]
    prices = {}
    for r in range(2, ws.max_row + 1):
        code = strip_str(ws.cell(r, 1).value)
        if not code: continue
        if code in prices: continue
        price = ws.cell(r, 2).value
        if isinstance(price, (int, float)):
            prices[code] = price
        elif price is not None and str(price).strip() != "":
            # Preserve text values like "POA"
            prices[code] = str(price).strip()
    return prices


def build_mini_products():
    print("→ Reading Mini Catalogue Self Updating + PartsbookBenji")
    catalogue_rows = read_mini_catalogue_codes()
    prices = read_partsbook_prices()
    out = []
    seen = set()
    next_id = 1
    poa_count = 0
    out_count = 0
    for code, desc, section in catalogue_rows:
        # De-dupe by (code, section) — same code can legitimately appear in
        # multiple sections (e.g. Apx1 + 120) if the catalogue lists it twice
        key = (code, section)
        if key in seen: continue
        seen.add(key)

        bt, mk, hd = parse_description(desc)
        price = prices.get(code)

        # Two possible states (real stock tracking comes in Phase 3):
        #   numeric price → in stock, show £
        #   anything else → in stock, show "POA" — covers "POA" text markers
        #                   AND parts in the catalogue that PartsbookBenji
        #                   doesn't price (customer should call to ask)
        if isinstance(price, (int, float)):
            price_ex = price
            price_inc = round2(price * 1.20)
        else:
            price_ex = None
            price_inc = None
            if price is not None and str(price).strip() != "":
                poa_count += 1
            else:
                out_count += 1
        stock = "in"  # always "in" until Phase 3 wires real stock tracking

        out.append({
            "id": "p" + str(next_id).zfill(4),
            "code": code,
            "name": desc,
            "section": section,
            "fits": fits_string(bt, mk),
            "bodyType": bt,
            "mark": mk,
            "hand": hd,
            "priceExVat": price_ex,
            "priceIncVat": price_inc,
            "stock": stock,
            "stockQty": 0,
            "category": "mini",
        })
        next_id += 1
    print(f"  ✓ Mini products: {len(out)}  ({poa_count} POA in master, {out_count} not in master — both shown as POA)")
    return out


# ─────────────────────────────────────────────────────────────────────────────
# Metals products — list from Metals catalogue 2023, prices from Metals.xlsx
# ─────────────────────────────────────────────────────────────────────────────

def build_metals_products():
    print("→ Reading Metals catalogue 2023 + Metals.xlsx")
    if not METALS_CAT.exists():
        print(f"  ! Metals catalogue not found at {METALS_CAT}; metals products will be empty")
        return []

    # Master lookup
    rows, _collisions = build_lookup_rows()
    keys, by_metal_size, by_metal_spec_size = build_indexes(rows)

    # Catalogue
    wb = openpyxl.load_workbook(METALS_CAT, data_only=True)
    out = []
    next_id = 1
    auto_linked = 0
    fallback = 0

    for sheet_name in wb.sheetnames:
        if sheet_name.startswith("_"): continue
        ws = wb[sheet_name]
        if not is_data_sheet(ws): continue

        cat = category_from_metal_sheet(sheet_name)

        for row_idx in range(2, ws.max_row + 1):
            shape = strip_str(ws.cell(row_idx, 1).value)
            metal = strip_str(ws.cell(row_idx, 2).value)
            spec  = strip_str(ws.cell(row_idx, 3).value)
            size  = strip_str(ws.cell(row_idx, 4).value)
            existing_price = ws.cell(row_idx, 5).value
            unit  = strip_str(ws.cell(row_idx, 6).value)

            if not size and not metal: continue

            matched_key, confidence = find_match(metal, spec, size, shape, keys, by_metal_size, by_metal_spec_size)

            price_ex = None
            poa = False
            if matched_key and confidence in HIGH_CONFIDENCE:
                master_price = keys[matched_key]["priceEx"]
                if isinstance(master_price, (int, float)):
                    price_ex = master_price
                else:
                    # Master has a text marker (POA or similar) — treat the
                    # row as in-stock with no numeric price
                    poa = True
                auto_linked += 1
            elif isinstance(existing_price, (int, float)):
                price_ex = existing_price
                fallback += 1
            elif existing_price is not None and str(existing_price).strip() != "":
                # Hardcoded text in catalogue (e.g. "POA")
                poa = True
                fallback += 1
            else:
                fallback += 1

            price_inc = round2(price_ex * 1.20) if price_ex is not None else None

            pieces = [shape, metal, spec, size]
            name = " — ".join(p for p in pieces if p)
            code = ((spec or metal or shape or sheet_name).replace(" ", "")[:32] +
                    ("-" + size.replace(" ", "") if size else ""))

            # Always "in" — real stock tracking comes in Phase 3. Pricing
            # state is communicated via priceExVat (null = "POA" badge in UI).
            stock = "in"

            out.append({
                "id": "m" + str(next_id).zfill(4),
                "code": code,
                "name": name,
                "category": cat,
                "form": shape,
                "stock": stock,
                "priceExVat": price_ex,
                "priceIncVat": price_inc,
                "unit": unit,
                "spec": spec,
                "size": size,
                "notes": "",
                "sourceSheet": sheet_name,
            })
            next_id += 1

    print(f"  ✓ Metals products: {len(out)}  (auto-linked {auto_linked}, hardcoded fallback {fallback})")
    return out


# ─────────────────────────────────────────────────────────────────────────────
# Featured Work — read JSON, copy images to public/, write lib/featured-data.ts
# ─────────────────────────────────────────────────────────────────────────────


def build_featured_work():
    print("→ Reading Featured Work JSON")
    if not FEATURED_JSON.exists():
        print(f"  ! featured-work.json not found at {FEATURED_JSON}; skipping")
        return None

    with open(FEATURED_JSON, "r", encoding="utf-8") as f:
        try:
            entries = json.load(f)
        except json.JSONDecodeError as e:
            print(f"  ! featured-work.json is not valid JSON: {e}")
            print(f"    Fix the JSON syntax and re-run.")
            sys.exit(1)

    if not isinstance(entries, list):
        print("  ! featured-work.json must be a JSON array (a list)")
        sys.exit(1)

    # Make sure public/featured exists
    FEATURED_IMG_DST.mkdir(parents=True, exist_ok=True)

    # Track which destination filenames we wrote so we can later prune ones
    # that were referenced previously but no longer are
    written = set()
    out = []
    missing_images = []

    for i, e in enumerate(entries):
        if not isinstance(e, dict):
            print(f"  ! entry {i} in featured-work.json is not an object — skipping")
            continue

        item = {
            "id":          str(e.get("id", f"f{i+1:03d}")),
            "title":       str(e.get("title", "Untitled")),
            "description": str(e.get("description", "")),
            "tag":         str(e.get("tag", "Bespoke")),
            "year":        int(e.get("year", datetime.now().year)),
            "category":    str(e.get("category", "Fabrication")),
            "fullStory":   str(e.get("fullStory", "")),
            "imagePath":   None,
        }

        img = (e.get("image") or "").strip()
        if img:
            src_img = FEATURED_IMG_SRC / img
            if not src_img.exists():
                missing_images.append((item["id"], img))
                # Still emit the entry — page will fall back to placeholder
            else:
                # Use a deterministic destination filename so the URL stays
                # stable even if the source name changes later. Keep the
                # original extension though.
                ext = src_img.suffix.lower()
                dst_name = f"{item['id']}{ext}"
                dst_path = FEATURED_IMG_DST / dst_name
                # Skip the copy if dst is identical to src
                try:
                    if not dst_path.exists() or dst_path.stat().st_size != src_img.stat().st_size:
                        shutil.copy2(src_img, dst_path)
                except PermissionError:
                    # Can't overwrite (maybe locked) — leave the existing copy
                    pass
                item["imagePath"] = f"/featured/{dst_name}"
                written.add(dst_name)

        out.append(item)

    # Tidy up old image files in public/featured/ that are no longer
    # referenced. We only sweep files matching the f###.<ext> pattern so we
    # don't accidentally delete unrelated files in that folder.
    for f in FEATURED_IMG_DST.iterdir():
        if f.is_file() and re.match(r"^f\d+\.[a-zA-Z]+$", f.name) and f.name not in written:
            try:
                f.unlink()
            except Exception:
                pass

    if missing_images:
        print(f"  ! {len(missing_images)} entries reference images that aren't in data-source/featured-images/:")
        for entry_id, fn in missing_images[:5]:
            print(f"     - {entry_id}: '{fn}' (drop this file in or update the JSON)")
        if len(missing_images) > 5:
            print(f"     ... plus {len(missing_images) - 5} more")

    print(f"  ✓ Featured work: {len(out)} entries  ({len(written)} with images, {len(missing_images)} missing images)")
    return out


def write_featured_data_file(entries):
    if entries is None:
        # File doesn't exist at all — leave the existing lib/featured-data.ts
        # untouched. (Avoids wiping the homepage if someone deletes the JSON.)
        print(f"  ↷ No featured-work.json — lib/featured-data.ts left as-is")
        return
    generated = datetime.utcnow().isoformat() + "Z"
    entries_json = json.dumps(entries, indent=2, ensure_ascii=False)

    content = f"""// AUTO-GENERATED by scripts/regen_website_data.py — DO NOT EDIT BY HAND
// Source:    data-source/featured-work.json (+ data-source/featured-images/)
// Generated: {generated}
// Re-run:    npm run sync-data

export type FeaturedWork = {{
  id: string;
  title: string;
  description: string;
  tag: string;
  year: number;
  category: string;
  fullStory: string;
  imagePath: string | null;  // null = no image, page falls back to placeholder
}};

export const featuredWork: FeaturedWork[] = {entries_json};
"""
    out_path = LIB / "featured-data.ts"
    out_path.write_text(content, encoding="utf-8")
    print(f"  ✓ Wrote lib/featured-data.ts ({len(entries)} entries)")


# ─────────────────────────────────────────────────────────────────────────────
# TypeScript writers
# ─────────────────────────────────────────────────────────────────────────────

def js_value(v):
    """Serialize a Python value to JSON. None → null."""
    return json.dumps(v, ensure_ascii=False)


def write_mini_data_file(products):
    generated = datetime.utcnow().isoformat() + "Z"
    sections_json = json.dumps(SECTIONS, indent=2, ensure_ascii=False)
    products_json = json.dumps(products, indent=2, ensure_ascii=False)

    content = f"""// AUTO-GENERATED by scripts/regen_website_data.py — DO NOT EDIT BY HAND
// Source:    data-source/Mini Catalogue Self Updating.xlsm  (product list)
// Prices:    data-source/PartsbookBenji2014.xlsx            (master)
// Generated: {generated}
// Re-run:    npm run sync-data

export type StockLevel = "in" | "low" | "out";
export type SectionMode = "exterior" | "interior";

export type Product = {{
  id: string;
  code: string;
  name: string;
  section: string;
  fits: string;
  bodyType: string | null;
  mark: string | null;
  hand: string | null;
  priceExVat: number | null;
  priceIncVat: number | null;
  stock: StockLevel;
  stockQty: number;
  category: "mini" | "metals";
}};

export type Section = {{
  code: string;
  label: string;
  subtitle: string;
  order: number;
  mode: SectionMode;
}};

export const sections: Section[] = {sections_json};

export const getSection = (code: string) => sections.find(s => s.code === code);

export const products: Product[] = {products_json};

export const getProductsBySection = (code: string) => products.filter(p => p.section === code);
export const getProductByCode = (code: string) => products.find(p => p.code === code);
"""
    out_path = LIB / "mini-data.ts"
    out_path.write_text(content, encoding="utf-8")
    print(f"  ✓ Wrote lib/mini-data.ts ({len(products)} products)")


def write_metals_data_file(metals):
    generated = datetime.utcnow().isoformat() + "Z"
    cats_json = json.dumps(METALS_CATEGORIES, indent=2, ensure_ascii=False)
    augmented = []
    for m in metals:
        m2 = dict(m)
        m2["pricePerKg"] = m["priceExVat"]
        m2["description"] = " — ".join(p for p in [m["spec"], m["notes"]] if p) or m["name"]
        augmented.append(m2)
    metals_json = json.dumps(augmented, indent=2, ensure_ascii=False)

    content = f"""// AUTO-GENERATED by scripts/regen_website_data.py — DO NOT EDIT BY HAND
// Source:    data-source/Metals catalogue 2023.xlsx  (product list)
// Prices:    data-source/Metals.xlsx                 (master)
// Generated: {generated}
// Re-run:    npm run sync-data

export type MetalProduct = {{
  id: string;
  code: string;
  name: string;
  category: string;
  form: string;
  stock: "in" | "low" | "out";
  priceExVat: number | null;
  priceIncVat: number | null;
  unit: string;
  spec: string;
  size: string;
  notes: string;
  sourceSheet: string;
  // legacy compat — prefer priceExVat / priceIncVat
  pricePerKg: number | null;
  description: string;
}};

export const metalCategories = {cats_json};

export const metals: MetalProduct[] = {metals_json};
"""
    out_path = LIB / "metals-data.ts"
    out_path.write_text(content, encoding="utf-8")
    print(f"  ✓ Wrote lib/metals-data.ts ({len(metals)} metals)")


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("\nM-Machine sync-data — regenerating website data from customer-facing docs\n")

    # Validate inputs
    missing = [str(p) for p in [PARTSBOOK, MINI_CAT, METALS_CAT, METALS] if not p.exists()]
    if missing:
        print("ERROR: missing required files in data-source/:")
        for m in missing:
            print(f"  - {m}")
        sys.exit(1)

    mini = build_mini_products()
    metals_list = build_metals_products()
    featured = build_featured_work()
    write_mini_data_file(mini)
    write_metals_data_file(metals_list)
    write_featured_data_file(featured)
    print("\n✓ Done. Refresh your browser to see the changes.\n")


if __name__ == "__main__":
    main()
