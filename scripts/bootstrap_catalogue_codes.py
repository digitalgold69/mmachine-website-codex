#!/usr/bin/env python3
"""One-time bootstrap of stable catalogue codes from the current Git version."""

import json
import subprocess
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl


HERE = Path(__file__).resolve().parent
PROJECT_ROOT = HERE.parent
sys.path.insert(0, str(HERE / "phase2"))

from catalogue_codes import save_codes  # noqa: E402
from metal_matching import make_catalogue_base, make_catalogue_link_id  # noqa: E402
from wire_catalogue import is_data_sheet, s  # noqa: E402


def read_old_products():
    source = subprocess.check_output(
        ["git", "show", "HEAD:lib/metals-data.ts"],
        cwd=PROJECT_ROOT,
        text=True,
        encoding="utf-8",
    )
    marker = "export const metals"
    start = source.index(marker)
    start = source.index("=", start) + 1
    start = source.index("[", start)
    return json.JSONDecoder().raw_decode(source[start:])[0]


def main():
    products = read_old_products()
    workbook = openpyxl.load_workbook(
        PROJECT_ROOT / "data-source" / "Metals catalogue 2023.xlsx",
        data_only=True,
    )
    codes = {}
    product_index = 0

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        if not is_data_sheet(ws):
            continue
        occurrences = defaultdict(int)
        for row in range(2, ws.max_row + 1):
            shape = s(ws.cell(row, 1).value)
            metal = s(ws.cell(row, 2).value)
            spec = s(ws.cell(row, 3).value)
            size = s(ws.cell(row, 4).value)
            unit = s(ws.cell(row, 6).value)
            if not size and not metal:
                continue
            if product_index >= len(products):
                raise RuntimeError("Catalogue has more rows than the current website data")
            base = make_catalogue_base(sheet_name, shape, metal, spec, size, unit)
            occurrences[base] += 1
            link_id = make_catalogue_link_id(
                sheet_name, shape, metal, spec, size, unit, occurrences[base]
            )
            codes[link_id] = products[product_index]["code"]
            product_index += 1

    if product_index != len(products):
        raise RuntimeError(
            f"Website has {len(products)} metals but catalogue supplied {product_index}"
        )
    save_codes(codes)
    print(f"Bootstrapped {len(codes)} stable public metals codes")


if __name__ == "__main__":
    main()
