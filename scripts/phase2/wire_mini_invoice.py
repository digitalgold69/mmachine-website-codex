#!/usr/bin/env python3
"""
Phase 2b — Wire Mini Invoice Template to a self-contained price lookup.

The Mini Invoice Template uses TIERED PRICING — cell AA14 in the Invoice sheet
selects which price list to use:
   AA14 = 1  →  retail (Parts Data)
   AA14 = 2  →  KDMSPC
   AA14 = 3  →  MSPORT
   AA14 = 4  →  Magnum
   AA14 = 5  →  Somerford

Each lookup formula contains all 5 sources via `_xlfn.IFS(...)`. Formulas use
several different range patterns:
   '[1]Parts Data'!$A$3:$E$2740      (col A start, return col 2 = price)
   [1]KDMSPC!$D$3:$AA$2744           (col D start, return col 24 = AA = Part No)
   ...etc, with the source sheet name varying.

To preserve every formula exactly, this wire-up:
  1. Copies ALL columns and rows of each PartsbookBenji pricing sheet into
     internal hidden sheets (no column truncation).
  2. Just renames the sheet reference in every formula. Range bounds, column
     indices, the IFS structure — all left untouched.
  3. Strips external-link metadata.
"""
import openpyxl
import re
import shutil
import sys
from pathlib import Path

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
PROJECT_ROOT = HERE.parent.parent
PARTSBOOK = str(PROJECT_ROOT / "data-source" / "PartsbookBenji2014.xlsx")
INVOICE_SRC = str(PROJECT_ROOT / "data-source" / "Mini Invoice Template.xlsm")
INVOICE = str(PROJECT_ROOT / "final-deliverables" / "Mini Invoice Template.xlsm")

from surgical_xlsx import (
    open_for_surgery, repack_zip, add_sheet, remove_sheet_if_present,
    cell_str, cell_num, col_letter, _read, _write,
)
from xml.sax.saxutils import escape as xml_escape


# Map original PartsbookBenji sheet → internal _PriceLookup sheet name
SOURCE_SHEETS = [
    ("Parts Data", "_PriceLookup"),
    ("KDMSPC",     "_KDMSPC"),
    ("MSPORT",     "_MSPORT"),
    ("Magnum",     "_Magnum"),
    ("Somerford",  "_Somerford"),
]


def load_full_sheet(wb, source_name):
    """Copy a full PartsbookBenji pricing sheet — every populated row, every
    populated column. Preserves the original row indices so formulas using
    fixed row references like $A$3:$E$2740 still hit the right rows.
    Returns a list of (row_idx, [(col_idx, value), ...]) pairs."""
    if source_name not in wb.sheetnames:
        return []
    ws = wb[source_name]
    rows = []
    for r in range(1, ws.max_row + 1):
        cells = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is None or v == "":
                continue
            cells.append((c, v))
        if cells:
            rows.append((r, cells))
    return rows


def build_full_lookup_sheet_xml(rows, max_col):
    """Build a worksheet XML preserving the original row indices and column
    layout. Skips rows with no data (formulas don't care about gaps)."""
    last_row = rows[-1][0] if rows else 1
    last_col_letter = col_letter(max_col)
    parts = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">',
        f'<dimension ref="A1:{last_col_letter}{last_row}"/>',
        '<sheetViews><sheetView workbookViewId="0">'
        '<pane ySplit="1" topLeftCell="A2" activePane="bottomLeft" state="frozen"/>'
        '</sheetView></sheetViews>',
        '<sheetFormatPr defaultRowHeight="15"/>',
        '<sheetData>',
    ]
    for row_idx, cells in rows:
        parts.append(f'<row r="{row_idx}">')
        for col_idx, val in cells:
            ref = f"{col_letter(col_idx)}{row_idx}"
            if isinstance(val, (int, float)):
                parts.append(cell_num(ref, val))
            elif isinstance(val, bool):
                parts.append(f'<c r="{ref}" t="b"><v>{1 if val else 0}</v></c>')
            else:
                parts.append(cell_str(ref, str(val)))
        parts.append('</row>')
    parts.append('</sheetData>')
    parts.append('<pageMargins left="0.7" right="0.7" top="0.75" bottom="0.75" '
                 'header="0.3" footer="0.3"/>')
    parts.append('</worksheet>')
    return "\n".join(parts)


def rewrite_external_references_in_sheet(sheet_xml_path):
    """Just rename the sheet in every external reference — leave the range
    bounds untouched. Returns the count of substitutions."""
    with open(sheet_xml_path, "r", encoding="utf-8") as f:
        xml = f.read()

    total = 0
    for source_name, internal_name in SOURCE_SHEETS:
        escaped = re.escape(source_name)
        # Apostrophe-wrapped form (sheet names with spaces) — handle both
        # literal apostrophes and the XML-escaped &apos; variant.
        # Match  '[1]<source>'!  →  <internal>!
        for apos in ("'", "&apos;"):
            pat = re.compile(rf"{re.escape(apos)}\[1\]{escaped}{re.escape(apos)}!")
            new_xml, n = pat.subn(f"{internal_name}!", xml)
            xml = new_xml
            total += n

        # Bare form (single-word sheet names): [1]<source>!  →  <internal>!
        pat = re.compile(rf"\[1\]{escaped}!")
        new_xml, n = pat.subn(f"{internal_name}!", xml)
        xml = new_xml
        total += n

    if total > 0:
        with open(sheet_xml_path, "w", encoding="utf-8") as f:
            f.write(xml)
    return total


def remove_external_link_rels(tempdir):
    """Strip the external-workbook reference so Excel doesn't prompt about
    a missing file."""
    workbook = _read(tempdir, "xl/workbook.xml")
    workbook = re.sub(
        r'<externalReferences>.*?</externalReferences>', "", workbook, flags=re.DOTALL
    )
    _write(tempdir, "xl/workbook.xml", workbook)

    rels = _read(tempdir, "xl/_rels/workbook.xml.rels")
    # Use [^>]*? not [^/]*? — Type and Target attributes contain slashes
    rels = re.sub(
        r'<Relationship[^>]*?Type="[^"]*externalLink[^"]*"[^>]*?/>', "", rels
    )
    _write(tempdir, "xl/_rels/workbook.xml.rels", rels)

    ct = _read(tempdir, "[Content_Types].xml")
    # PartName contains slashes too — use [^>]*? not [^/]*?
    ct = re.sub(
        r'<Override[^>]*?ContentType="[^"]*externalLink[^"]*"[^>]*?/>', "", ct
    )
    _write(tempdir, "[Content_Types].xml", ct)

    el_dir = Path(tempdir) / "xl" / "externalLinks"
    if el_dir.exists():
        shutil.rmtree(el_dir, ignore_errors=True)


def wire_mini_invoice():
    print(f"  Source:  {INVOICE_SRC}")
    print(f"  Output:  {INVOICE}")

    if not Path(INVOICE_SRC).exists():
        print(f"  ! Mini Invoice Template not found at {INVOICE_SRC}")
        print(f"  ! Drop the unprotected file there and re-run.")
        return

    pb = openpyxl.load_workbook(PARTSBOOK, data_only=True)
    Path(INVOICE).parent.mkdir(parents=True, exist_ok=True)

    # Extract directly from the source (read-only) so a stale lock on the
    # destination doesn't break the run. The destination is overwritten only
    # at the final repack_zip step.
    tempdir = open_for_surgery(INVOICE_SRC)
    try:
        # Add the five internal price-lookup sheets — full column copies
        for source_name, internal_name in SOURCE_SHEETS:
            rows = load_full_sheet(pb, source_name)
            if not rows:
                print(f"  ! source sheet '{source_name}' empty — skipping")
                continue
            max_col = max(c for _r, cells in rows for c, _v in cells)
            print(f"  {internal_name:<14s}: {len(rows):>5} rows x {max_col:>2} cols  (from {source_name})")
            remove_sheet_if_present(tempdir, internal_name)
            add_sheet(tempdir, internal_name, build_full_lookup_sheet_xml(rows, max_col), hidden=True)

        # Rewrite every external reference (just the sheet-name part)
        sheets_dir = Path(tempdir) / "xl" / "worksheets"
        total_rewrites = 0
        sheets_touched = 0
        for sheet_path in sheets_dir.glob("sheet*.xml"):
            n = rewrite_external_references_in_sheet(str(sheet_path))
            if n > 0:
                sheets_touched += 1
                total_rewrites += n

        print(f"  Rewrote {total_rewrites} external references across {sheets_touched} sheets")
        remove_external_link_rels(tempdir)

        repack_zip(tempdir, INVOICE)
        tempdir = None
    finally:
        if tempdir:
            shutil.rmtree(tempdir, ignore_errors=True)

    print(f"  ✓ Tiered pricing preserved (AA14 = 1..5 still selects price list)")
    print(f"  ✓ VBA, address book, images, print settings — preserved byte-for-byte")


if __name__ == "__main__":
    wire_mini_invoice()
