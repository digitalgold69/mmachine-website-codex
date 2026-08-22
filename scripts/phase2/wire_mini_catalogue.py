#!/usr/bin/env python3
"""
Phase 2b - Refresh the Mini customer catalogue from Partsbook prices.

The Mini Catalogue currently uses external-link VLOOKUPs into
PartsbookBenji2014.xlsx — that's the fragile pattern that breaks when the
files move or get emailed around. The generated customer catalogue instead
contains current displayed prices and no external workbook dependency.

Strategy:
  1. Read PartsbookBenji2014.xlsx → 'Parts Data'. Extract (code, price) rows.
  2. Write current ex-VAT and inc-VAT values into the existing price cells.
  3. Remove the old external link and stale calculation chain.
  4. Repack as .xlsm (preserve VBA, images, print settings, etc.)

The output does not add worksheets or alter the workbook relationship layout,
which keeps this old macro workbook compatible with Excel 2007.
"""
import openpyxl
import json
import re
import shutil
import sys
import warnings
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(errors="replace")

warnings.filterwarnings("ignore", message="DrawingML support is incomplete.*")
warnings.filterwarnings("ignore", message="Print area cannot be set.*")
warnings.filterwarnings("ignore", message="Cannot parse header or footer.*")
warnings.filterwarnings("ignore", message="Data Validation extension is not supported.*")

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
PROJECT_ROOT = HERE.parent.parent
from source_paths import source_file

PARTSBOOK = str(source_file("PartsbookBenji2014.xlsx"))
MINI_CAT_SRC = str(source_file("Mini Catalogue Self Updating.xlsm"))
MINI_CAT = str(PROJECT_ROOT / "final-deliverables" / "Mini Catalogue Self Updating.xlsm")
MINI_MANIFEST = str(PROJECT_ROOT / "final-deliverables" / "mini-catalogue-updates.json")

from surgical_xlsx import (
    open_for_surgery, repack_zip, remove_sheet_if_present,
    remove_calc_chain_if_present, map_sheet_names_to_xml_files, edit_sheet_xml,
    cell_str, cell_num, col_letter, _read, _write,
)


PART_CODE_REGEX = re.compile(r"^[\d.]+\.\d{2}\.\d{2}\.\d{2}[A-Z]?$|^[A-Z0-9-]{4,}$")


def looks_like_part_code(value):
    value = str(value or "").strip()
    return (
        bool(value)
        and " " not in value
        and 3 <= len(value) <= 30
        and bool(PART_CODE_REGEX.match(value))
    )


def code_description_columns(ws):
    pairs = []
    for col in range(1, ws.max_column + 1):
        if str(ws.cell(1, col).value or "").strip().lower() == "code":
            pairs.append((col, col + 1))
    return pairs


def value_cell(ref, value):
    if isinstance(value, (int, float)):
        return cell_num(ref, value)
    return cell_str(ref, " " if value is None else str(value))


def build_price_value_edits(ws, prices, manifest_updates):
    edits = {}
    for code_col, _description_col in code_description_columns(ws):
        price_col = code_col + 2
        inc_vat_col = code_col + 3
        for row in range(2, ws.max_row + 1):
            raw_code = ws.cell(row, code_col).value
            if not looks_like_part_code(raw_code):
                continue
            code = str(raw_code).strip()
            price_ref = f"{col_letter(price_col)}{row}"
            inc_ref = f"{col_letter(inc_vat_col)}{row}"
            current_price = prices.get(code)
            cached_price = current_price if current_price is not None else " "
            cached_inc_vat = (
                round(float(current_price) * 1.2, 2)
                if isinstance(current_price, (int, float))
                else " "
            )
            edits.setdefault(row, []).append(
                (price_ref, value_cell(price_ref, cached_price))
            )
            manifest_updates.append(
                {"sheet": ws.title, "cell": price_ref, "value": cached_price}
            )
            edits.setdefault(row, []).append(
                (inc_ref, value_cell(inc_ref, cached_inc_vat))
            )
            manifest_updates.append(
                {"sheet": ws.title, "cell": inc_ref, "value": cached_inc_vat}
            )

    # The legacy catalogue pre-fills some empty future rows with external
    # VLOOKUP formulas even though their part-code cell is blank. They are not
    # products, but leaving those formulas behind keeps a broken external
    # workbook dependency. Blank only those unused lookup cells.
    edited_refs = {
        ref
        for row_edits in edits.values()
        for ref, _cell_xml in row_edits
    }
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if (
                cell.coordinate not in edited_refs
                and isinstance(value, str)
                and value.startswith("=")
                and "[1]Parts Data" in value
            ):
                edits.setdefault(cell.row, []).append(
                    (cell.coordinate, value_cell(cell.coordinate, " "))
                )
                manifest_updates.append(
                    {"sheet": ws.title, "cell": cell.coordinate, "value": " "}
                )
    return edits


def load_partsbook_codes():
    """Read PartsbookBenji.xlsx → 'Parts Data' and return a list of
    (code, price) pairs. Price may be numeric or a text marker like 'POA'.
    Skips heading rows and duplicates."""
    wb = openpyxl.load_workbook(PARTSBOOK, data_only=True)
    ws = wb["Parts Data"]
    rows = []
    seen = set()
    for r in range(2, ws.max_row + 1):
        code = ws.cell(r, 1).value
        price = ws.cell(r, 2).value
        if code is None: continue
        code_s = str(code).strip()
        if not code_s: continue
        if code_s in seen: continue   # PartsbookBenji has a few duplicates
        seen.add(code_s)
        # Preserve text values like "POA" (Price on application) — they
        # used to flow through the original external VLOOKUP and we want
        # the catalogue to behave the same way.
        rows.append((code_s, price))
    return rows


def remove_external_link_rels(tempdir):
    """Strip the external-workbook reference from the workbook so Excel
    doesn't prompt about a missing file. The reference lives in:
       xl/workbook.xml         (<externalReferences>)
       xl/_rels/workbook.xml.rels
       xl/externalLinks/...    (the actual cached reference data)
    """
    # workbook.xml — drop the <externalReferences>...</externalReferences>
    # block AND any <externalReference r:id="rIdN"/> entries inside.
    workbook = _read(tempdir, "xl/workbook.xml")
    workbook = re.sub(
        r'<externalReferences>.*?</externalReferences>', "", workbook, flags=re.DOTALL
    )
    _write(tempdir, "xl/workbook.xml", workbook)

    # workbook.xml.rels — drop external-link Relationships
    # Use [^>]*? not [^/]*? — Type and Target attribute values contain slashes
    rels = _read(tempdir, "xl/_rels/workbook.xml.rels")
    rels = re.sub(
        r'<Relationship[^>]*?Type="[^"]*externalLink[^"]*"[^>]*?/>', "", rels
    )
    _write(tempdir, "xl/_rels/workbook.xml.rels", rels)

    # [Content_Types].xml — drop the externalLink override (PartName has slashes)
    ct = _read(tempdir, "[Content_Types].xml")
    ct = re.sub(
        r'<Override[^>]*?ContentType="[^"]*externalLink[^"]*"[^>]*?/>', "", ct
    )
    _write(tempdir, "[Content_Types].xml", ct)

    # And remove the externalLinks/ folder content (best-effort)
    el_dir = Path(tempdir) / "xl" / "externalLinks"
    if el_dir.exists():
        shutil.rmtree(el_dir, ignore_errors=True)
    el_rels_dir = Path(tempdir) / "xl" / "externalLinks" / "_rels"
    if el_rels_dir.exists():
        shutil.rmtree(el_rels_dir, ignore_errors=True)


def wire_mini_catalogue():
    print(f"  Source:  {MINI_CAT_SRC}")
    print(f"  Output:  {MINI_CAT}")

    rows = load_partsbook_codes()
    prices = dict(rows)
    print(f"  PartsbookBenji codes: {len(rows)}")
    source_wb = openpyxl.load_workbook(MINI_CAT_SRC, data_only=False, keep_vba=True)
    manifest_updates = []

    Path(MINI_CAT).parent.mkdir(parents=True, exist_ok=True)
    # Extract directly from the source (avoids permission error if destination is open)
    tempdir = open_for_surgery(MINI_CAT_SRC)
    try:
        # A previous generated file should never be used as the source, but
        # remove the old helper sheet if one is present so the build stays
        # idempotent.
        remove_sheet_if_present(tempdir, "_PriceLookup")

        # The customer catalogue is an output document. Embed current values in
        # its existing cells instead of adding a worksheet and live formulas.
        # This preserves the old workbook structure required by Excel 2007.
        sheet_path_map = map_sheet_names_to_xml_files(tempdir)
        total_updates = 0
        sheets_touched = 0
        for sheet_name, xml_rel in sheet_path_map.items():
            if not (re.fullmatch(r"\d+B", sheet_name) or sheet_name in {"APX1", "APX2"}):
                continue
            sheet_path = Path(tempdir) / xml_rel
            edits = build_price_value_edits(
                source_wb[sheet_name], prices, manifest_updates
            )
            if edits:
                _applied, missing = edit_sheet_xml(str(sheet_path), edits)
                if missing:
                    raise RuntimeError(
                        f"{sheet_name}: could not update price rows {missing[:12]}"
                    )
            inserted = sum(len(row_edits) for row_edits in edits.values())
            if inserted:
                sheets_touched += 1
                total_updates += inserted

        print(
            f"  Embedded {total_updates} current price values across "
            f"{sheets_touched} catalogue pages"
        )

        # Strip the external-link metadata so Excel doesn't prompt
        remove_external_link_rels(tempdir)

        # Price formulas changed to values, so the source dependency chain is
        # stale. Let Excel generate a fresh chain if the output is ever edited.
        remove_calc_chain_if_present(tempdir)

        repack_zip(tempdir, MINI_CAT)
        tempdir = None

        Path(MINI_MANIFEST).write_text(
            json.dumps(
                {
                    "source": Path(MINI_CAT_SRC).name,
                    "customerFile": Path(MINI_CAT).name,
                    "updates": manifest_updates,
                },
                ensure_ascii=True,
                separators=(",", ":"),
            ),
            encoding="utf-8",
        )
    finally:
        source_wb.close()
        if tempdir:
            shutil.rmtree(tempdir, ignore_errors=True)

    print("  OK Current Partsbook prices embedded without adding worksheets")
    print(f"  OK Excel update manifest written ({len(manifest_updates)} cells)")
    print("  OK VBA, images, and print settings preserved")


if __name__ == "__main__":
    wire_mini_catalogue()
