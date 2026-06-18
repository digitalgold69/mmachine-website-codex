#!/usr/bin/env python3
"""
Phase 2 Step 1 — Build the _PriceLookup helper sheet inside Metals.xlsx.

Reads every metal-data sheet in Metals.xlsx (master), normalises the rows into
a single flat table with a composite key, and writes that table to a hidden
sheet called `_PriceLookup` at the back of the workbook.

The composite key is:   <Metal>|<NormalisedSpec>|<NormalisedSize>|<NormalisedShape>
NormalisedSize and NormalisedSpec strip whitespace, lowercase, replace fancy
quotes with regular quotes, and collapse multiple spaces.

Every row in _PriceLookup has columns:
   Key  |  Shape  |  Metal  |  Spec  |  Size  |  £ ex VAT  |  Unit  |  Source Sheet  |  Source Row

The catalogue and invoice can VLOOKUP against the Key column.
"""
import openpyxl
import re
import sys
import os
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment
from metal_matching import (
    assign_master_link_ids,
    canonical_shape,
    is_shape_heading,
)

# Resolve paths relative to the project root (parent of scripts/phase2/)
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
MASTER = str(PROJECT_ROOT / "data-source" / "Metals.xlsx")

def s(v):
    if v is None: return ""
    return str(v).strip()

def n(v):
    if v is None or v == "": return None
    if isinstance(v, (int, float)): return float(v)
    try:
        return float(re.sub(r"[£,\s]", "", str(v)))
    except ValueError:
        return None

def normalise(v):
    """Lowercase, strip, normalise quotes and whitespace."""
    if not v: return ""
    s = str(v).strip().lower()
    s = s.replace("\u201c", '"').replace("\u201d", '"').replace("\u2018", "'").replace("\u2019", "'")
    s = s.replace("\u00a0", " ")          # nbsp -> space
    s = re.sub(r"\s+", " ", s)
    return s

# Sheet name → metal type
def metal_from_sheet(name):
    n = name.lower()
    if n.startswith("alu"): return "Aluminium"
    if n.startswith("brass"): return "Brass"
    if n.startswith("ph brnz"): return "Phosphor Bronze"
    if n.startswith("cu & nil ag"): return "Cupro-Nickel"
    if n.startswith("cu"): return "Copper"
    if n.startswith("st steel"): return "Stainless Steel"
    if n.startswith("steels"): return "Mild Steel"
    if n.startswith("cast i"): return "Cast Iron"
    if n.startswith("plastics"): return "Plastics"
    if n.startswith("misc"): return "Misc"
    return None

def parse_master():
    wb = openpyxl.load_workbook(MASTER, data_only=True)
    rows = []
    for sheet_name in wb.sheetnames:
        metal = metal_from_sheet(sheet_name)
        if not metal: continue
        ws = wb[sheet_name]
        # Header row 1: Shape | Metal | Spec. | Size | £ ex VAT | Unit | £ Inc VAT | Notes ...
        current_shape = ""
        current_grade = ""
        current_metal = ""
        legacy_shape = ""
        for row_idx in range(2, ws.max_row + 1):
            column_a = s(ws.cell(row_idx, 1).value)
            metal_cell = s(ws.cell(row_idx, 2).value)
            spec  = s(ws.cell(row_idx, 3).value)
            size  = s(ws.cell(row_idx, 4).value)
            priceEx = n(ws.cell(row_idx, 5).value)
            unit  = s(ws.cell(row_idx, 6).value)

            if not any((column_a, metal_cell, spec, size, priceEx is not None, unit)):
                current_grade = ""
                current_metal = ""
                continue

            legacy_row_shape = column_a or legacy_shape

            # Column A is overloaded. Shape-only rows are section headings,
            # while values on priced rows are normally grades/alloys such as
            # 6082T6, CZ126, or 1.4305.
            if column_a and not metal_cell and not spec and not size and priceEx is None:
                if is_shape_heading(column_a):
                    current_shape = canonical_shape(column_a)
                    current_grade = ""
                    current_metal = ""
                else:
                    current_grade = column_a
                legacy_shape = column_a
                continue

            if not metal_cell and not size: continue
            if priceEx is None: continue

            row_shape = canonical_shape(column_a) if is_shape_heading(column_a) else ""
            if row_shape:
                current_shape = row_shape
                current_grade = ""
            elif column_a:
                current_grade = column_a

            if metal_cell:
                current_metal = metal_cell

            rows.append({
                "shape": row_shape or current_shape,
                "legacyShape": legacy_row_shape,
                "metal": metal_cell or current_metal or metal,
                "legacyMetal": metal_cell or metal,
                "grade": current_grade,
                "spec": spec,
                "size": size,
                "priceEx": priceEx,
                "unit": unit,
                "sourceSheet": sheet_name,
                "sourceRow": row_idx,
            })
    assign_master_link_ids(rows)
    return rows

def make_key(metal, spec, size, shape=""):
    return "|".join([
        normalise(metal),
        normalise(spec),
        normalise(size),
        normalise(shape),
    ])

def build_lookup_rows():
    """Pure in-memory build. Returns a list of 12-tuples.
    (Key, Shape, Metal, Spec, Size, £ ex VAT, Unit, Source Sheet, Source Row, Code).

    Used by wire_catalogue.py, wire_invoice.py, regen_website_data.py.
    Codes are assigned via metal_codes.assign_codes() and persisted to
    data-source/.metal-codes.json so they're stable across syncs.
    """
    import sys as _sys
    _here = Path(__file__).resolve().parent
    if str(_here) not in _sys.path:
        _sys.path.insert(0, str(_here))
    from metal_codes import (
        assign_codes, composite_key, load_codes_file, save_codes_file
    )

    rows = parse_master()
    out = []
    seen = {}
    collisions = 0

    # Build the input for the code assigner (one entry per master row)
    code_inputs = [
        {
            "metal": r["legacyMetal"],
            "spec": r["spec"],
            "size": r["size"],
            "shape": r["legacyShape"],
        }
        for r in rows
    ]

    # Load existing code mapping (preserves codes across syncs)
    codes_path = PROJECT_ROOT / "data-source" / ".metal-codes.json"
    existing = load_codes_file(codes_path)
    code_map = assign_codes(code_inputs, existing=existing)

    # Persist any newly-assigned codes back to disk
    if code_map != existing:
        save_codes_file(codes_path, code_map)

    for r in rows:
        k = make_key(r["metal"], r["spec"], r["size"], r["shape"])
        if k in seen:
            collisions += 1
            k = k + "#" + r["sourceSheet"] + ":" + str(r["sourceRow"])
        seen[k] = r
        ckey = composite_key(
            r["legacyMetal"], r["spec"], r["size"], r["legacyShape"]
        )
        code = code_map.get(ckey, "")
        out.append((k, r["shape"], r["metal"], r["spec"], r["size"],
                    r["priceEx"], r["unit"], r["sourceSheet"], r["sourceRow"],
                    code, r["grade"], r["linkId"]))
    return out, collisions

def main():
    print(f"Reading {MASTER}")
    rows, collisions = build_lookup_rows()
    print(f"  Built _PriceLookup in memory: {len(rows)} rows ({collisions} collision keys disambiguated)")
    coded = sum(1 for r in rows if r[9])
    print(f"  Codes assigned: {coded}/{len(rows)}")

if __name__ == "__main__":
    main()
