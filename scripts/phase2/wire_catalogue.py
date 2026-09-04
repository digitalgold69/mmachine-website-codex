#!/usr/bin/env python3
"""
Phase 2 — Wire Metals catalogue 2023 to a self-contained _PriceLookup.

Surgical-edit version:
  • We use openpyxl ONLY for reading the catalogue (to figure out which
    rows match the master). We never let openpyxl save it.
  • All writes happen at the zip-level: we extract the original .xlsx,
    edit only the specific cells we need to change, add the two new sheets,
    and re-zip. Print settings, images, headers/footers, exotic page-layout
    settings — everything we don't touch is preserved byte-for-byte.

Pattern:
  • Master Metals.xlsx is the source of truth (read-only).
  • Catalogue gets its OWN copy of _PriceLookup (hidden sheet inside it).
  • Catalogue's E column does VLOOKUP against ITS OWN _PriceLookup.
  • Each sync run rebuilds the catalogue from scratch with fresh prices.
  • Rows that no longer match Metals.xlsx are shown as POA, never with an old
    catalogue workbook price.
"""
import openpyxl
import re
import shutil
import sys
import warnings
from collections import defaultdict
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(errors="replace")

warnings.filterwarnings("ignore", message="Print area cannot be set.*")
warnings.filterwarnings("ignore", message="Cannot parse header or footer.*")
warnings.filterwarnings("ignore", message="Data Validation extension is not supported.*")

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
PROJECT_ROOT = HERE.parent.parent
from source_paths import source_file

CATALOGUE_SRC = str(source_file("Metals catalogue 2023.xlsx"))
CATALOGUE = str(PROJECT_ROOT / "final-deliverables" / "Metals catalogue 2023.xlsx")

from build_lookup import build_lookup_rows
from catalogue_codes import CatalogueCodeRegistry
from metal_matching import (
    MetalMatcher,
    make_catalogue_base,
    make_catalogue_link_id,
)
from surgical_xlsx import (
    open_for_surgery, repack_zip, add_sheet, remove_sheet_if_present,
    remove_calc_chain_if_present,
    map_sheet_names_to_xml_files, edit_sheet_xml, extend_sheet_dimension,
    cell_str, cell_num, cell_formula, col_letter, hide_column_in_sheet,
)


def s(v):
    if v is None: return ""
    return str(v).strip()


def build_indexes(rows):
    """Build a direct lookup-key index for catalogue formulas."""
    keys = {}
    for row in rows:
        key, shape, metal, spec, size, priceEx, unit, srcSheet, srcRow = row[:9]
        keys[key] = {
            "shape": shape, "metal": metal, "spec": spec, "size": size,
            "priceEx": priceEx, "unit": unit, "src": f"{srcSheet}!{srcRow}",
            "stockSize": row[12] if len(row) > 12 else "",
        }
    return keys


def is_data_sheet(ws):
    a1 = ws.cell(1, 1).value
    return a1 and "shape" in str(a1).lower()


def build_pricelookup_xml(rows):
    """Build the worksheet XML for the hidden _PriceLookup sheet.
    Layout: Key | Shape | Metal | Spec | Size | £ ex VAT | Unit |
            Source Sheet | Source Row | Code
    """
    headers = ["Key", "Shape", "Metal", "Spec", "Size", "£ ex VAT", "Unit",
               "Source Sheet", "Source Row", "Code", "Grade", "Stable Link ID", "Stock length"]
    n_rows = len(rows) + 1
    last_col = col_letter(len(headers))

    parts = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">',
        f'<dimension ref="A1:{last_col}{n_rows}"/>',
        '<sheetViews><sheetView workbookViewId="0">'
        '<pane ySplit="1" topLeftCell="A2" activePane="bottomLeft" state="frozen"/>'
        '</sheetView></sheetViews>',
        '<sheetFormatPr defaultRowHeight="15"/>',
        '<cols>',
        '<col min="1" max="1" width="70" customWidth="1"/>',
        '<col min="2" max="4" width="16" customWidth="1"/>',
        '<col min="5" max="5" width="22" customWidth="1"/>',
        '<col min="6" max="6" width="10" customWidth="1"/>',
        '<col min="7" max="9" width="16" customWidth="1"/>',
        '<col min="10" max="10" width="32" customWidth="1"/>',
        '<col min="11" max="13" width="42" customWidth="1"/>',
        '</cols>',
        '<sheetData>',
    ]

    parts.append('<row r="1">')
    for ci, h in enumerate(headers, 1):
        parts.append(cell_str(f"{col_letter(ci)}1", h))
    parts.append('</row>')

    for ri, row in enumerate(rows, start=2):
        parts.append(f'<row r="{ri}">')
        # Includes the stable matching data plus Metals.xlsx column L stock length.
        for ci, val in enumerate(row[:13], 1):
            ref = f"{col_letter(ci)}{ri}"
            if val is None or val == "":
                continue
            if isinstance(val, (int, float)):
                parts.append(cell_num(ref, val))
            else:
                parts.append(cell_str(ref, val))
        parts.append('</row>')

    parts.append('</sheetData>')
    parts.append('<pageMargins left="0.7" right="0.7" top="0.75" bottom="0.75" '
                 'header="0.3" footer="0.3"/>')
    parts.append('</worksheet>')
    return "\n".join(parts)


def build_reviewme_xml(review_rows):
    """Build the worksheet XML for the visible _ReviewMe sheet."""
    headers = ["Sheet", "Row", "Shape", "Metal", "Spec", "Size",
               "Existing E value", "Matched Key", "Confidence", "Matched Master Row"]
    n_rows = len(review_rows) + 1
    last_col = col_letter(len(headers))

    parts = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">',
        f'<dimension ref="A1:{last_col}{n_rows}"/>',
        '<sheetViews><sheetView workbookViewId="0">'
        '<pane ySplit="1" topLeftCell="A2" activePane="bottomLeft" state="frozen"/>'
        '</sheetView></sheetViews>',
        '<sheetFormatPr defaultRowHeight="15"/>',
        '<cols>',
        '<col min="1" max="1" width="22" customWidth="1"/>',
        '<col min="2" max="2" width="6" customWidth="1"/>',
        '<col min="3" max="3" width="14" customWidth="1"/>',
        '<col min="4" max="4" width="12" customWidth="1"/>',
        '<col min="5" max="5" width="18" customWidth="1"/>',
        '<col min="6" max="6" width="22" customWidth="1"/>',
        '<col min="7" max="7" width="14" customWidth="1"/>',
        '<col min="8" max="8" width="60" customWidth="1"/>',
        '<col min="9" max="9" width="14" customWidth="1"/>',
        '<col min="10" max="10" width="22" customWidth="1"/>',
        '</cols>',
        '<sheetData>',
    ]

    parts.append('<row r="1">')
    for ci, h in enumerate(headers, 1):
        parts.append(cell_str(f"{col_letter(ci)}1", h))
    parts.append('</row>')

    for ri, r in enumerate(review_rows, start=2):
        parts.append(f'<row r="{ri}">')
        cells = [r["sheet"], r["row"], r["shape"], r["metal"], r["spec"], r["size"],
                 r["previousPrice"], r["matchedKey"], r["confidence"], r["matchedSrc"]]
        for ci, val in enumerate(cells, 1):
            ref = f"{col_letter(ci)}{ri}"
            if val is None or val == "":
                continue
            if isinstance(val, (int, float)):
                parts.append(cell_num(ref, val))
            else:
                parts.append(cell_str(ref, val))
        parts.append('</row>')

    parts.append('</sheetData>')
    parts.append(f'<autoFilter ref="A1:{last_col}{n_rows}"/>')
    parts.append('<pageMargins left="0.7" right="0.7" top="0.75" bottom="0.75" '
                 'header="0.3" footer="0.3"/>')
    parts.append('</worksheet>')
    return "\n".join(parts)


def wire_catalogue():
    print(f"  Source:  {CATALOGUE_SRC}")
    print(f"  Output:  {CATALOGUE}")

    rows, _collisions = build_lookup_rows()
    keys = build_indexes(rows)
    matcher = MetalMatcher(rows)
    code_registry = CatalogueCodeRegistry()
    print(f"  Master rows: {len(rows)}")

    # --- 1. Read catalogue with openpyxl (do NOT save) -------------------
    src_wb = openpyxl.load_workbook(CATALOGUE_SRC)
    sheet_edits = {}    # sheet_name -> {row_idx: [(ref, cell_xml), ...]}
    review_rows = []
    counts = defaultdict(int)

    for sheet_name in src_wb.sheetnames:
        if sheet_name.startswith("_"):
            continue
        ws = src_wb[sheet_name]
        if not is_data_sheet(ws):
            continue

        edits = {}
        occurrences = defaultdict(int)

        for row_idx in range(2, ws.max_row + 1):
            shape = s(ws.cell(row_idx, 1).value)
            metal = s(ws.cell(row_idx, 2).value)
            spec  = s(ws.cell(row_idx, 3).value)
            size  = s(ws.cell(row_idx, 4).value)
            existing_price = ws.cell(row_idx, 5).value
            unit = s(ws.cell(row_idx, 6).value)
            row_edits = []

            # Always rewrite legacy inc-VAT formulas before row skipping.
            # The source workbook uses shared formulas in column G; replacing
            # only some rows can leave Excel with broken shared-formula groups.
            g_val = ws.cell(row_idx, 7).value
            if isinstance(g_val, str) and g_val.startswith("="):
                g_formula = f'IF(E{row_idx}="","",E{row_idx}*1.2)'
                row_edits.append((f"G{row_idx}", cell_formula(f"G{row_idx}", g_formula)))

            if not size and not metal:
                counts["no-data"] += 1
                if row_edits:
                    edits[row_idx] = row_edits
                continue

            base = make_catalogue_base(
                sheet_name, shape, metal, spec, size, unit
            )
            occurrences[base] += 1
            link_id = make_catalogue_link_id(
                sheet_name, shape, metal, spec, size, unit, occurrences[base]
            )
            result = matcher.match(
                {
                    "shape": shape,
                    "metal": metal,
                    "spec": spec,
                    "size": size,
                    "unit": unit,
                },
                link_id,
                existing_price=existing_price,
            )
            matched_key = result.lookup_key
            confidence = result.confidence
            counts[confidence] += 1
            public_code = code_registry.get(
                link_id, metal, spec, size, shape
            )

            if matched_key:
                # E (price) looks up via VLOOKUP. H (Code) does the same.
                # The catalogue's _PriceLookup layout:
                #   A=Key, B=Shape, C=Metal, D=Spec, E=Size, F=£ ex VAT, G=Unit,
                #   H=Source Sheet, I=Source Row, J=Code
                e_formula = f'IFERROR(VLOOKUP(K{row_idx},_PriceLookup!$A:$F,6,FALSE),0)'
                current_price = keys[matched_key]["priceEx"]
                row_edits.append(
                    (
                        f"E{row_idx}",
                        cell_formula(f"E{row_idx}", e_formula, current_price),
                    )
                )
                g_formula = f'IF(E{row_idx}="","",E{row_idx}*1.2)'
                current_inc_vat = (
                    round(float(current_price) * 1.2, 2)
                    if isinstance(current_price, (int, float))
                    else ""
                )
                row_edits = [
                    edit for edit in row_edits if edit[0] != f"G{row_idx}"
                ]
                row_edits.append(
                    (
                        f"G{row_idx}",
                        cell_formula(f"G{row_idx}", g_formula, current_inc_vat),
                    )
                )
                row_edits.append((f"H{row_idx}", cell_str(f"H{row_idx}", public_code)))
                row_edits.append((f"K{row_idx}", cell_str(f"K{row_idx}", matched_key)))
                counts["auto_linked"] += 1
            else:
                # Not auto-linked — but still want a Code in column H so the
                # owner / customers can read it off the catalogue. Use the
                # candidate code from this row's own text (no master lookup).
                row_edits = [
                    edit for edit in row_edits if edit[0] != f"G{row_idx}"
                ]
                row_edits.append((f"E{row_idx}", cell_str(f"E{row_idx}", "POA")))
                row_edits.append((f"G{row_idx}", cell_str(f"G{row_idx}", "POA")))
                if public_code:
                    row_edits.append((f"H{row_idx}", cell_str(f"H{row_idx}", public_code)))
                review_rows.append({
                    "sheet": sheet_name, "row": row_idx,
                    "shape": shape, "metal": metal, "spec": spec, "size": size,
                    "previousPrice": existing_price,
                    "matchedKey": matched_key if matched_key else "(no match)",
                    "confidence": confidence,
                    "matchedSrc": keys[matched_key]["src"] if matched_key else "",
                })

            if row_edits:
                edits[row_idx] = row_edits

        if edits:
            sheet_edits[sheet_name] = edits

    src_wb.close()
    matcher.save()
    code_registry.save()

    review_rows.sort(key=lambda r: (r["confidence"], r["sheet"], r["row"]))

    # --- 2. Surgical edits — extract directly from the source ------------
    # (avoids a permission error if the destination is open in Excel)
    Path(CATALOGUE).parent.mkdir(parents=True, exist_ok=True)
    tempdir = open_for_surgery(CATALOGUE_SRC)
    try:
        # Be idempotent
        for cleanup in ("_PriceLookup", "_ReviewMe"):
            remove_sheet_if_present(tempdir, cleanup)
        remove_calc_chain_if_present(tempdir)

        # Apply cell edits to each metal data sheet
        sheet_path_map = map_sheet_names_to_xml_files(tempdir)
        for sheet_name, edits in sheet_edits.items():
            xml_rel = sheet_path_map.get(sheet_name)
            if not xml_rel:
                print(f"  ! sheet '{sheet_name}' not found in workbook xml - skipping")
                continue
            sheet_path = Path(tempdir) / xml_rel

            # Add a "Code" header at H1 (next to £ Inc VAT) and a "Lookup Key"
            # header at K1 (the existing hidden helper). H1 is visible — that's
            # the column the owner shows customers.
            if edits:
                edits.setdefault(1, []).append(("H1", cell_str("H1", "Code")))
                edits.setdefault(1, []).append(("K1", cell_str("K1", "Lookup Key")))

            applied, missing = edit_sheet_xml(str(sheet_path), edits)
            if missing:
                print(
                    f"  WARNING: {sheet_name} has {len(missing)} XML rows "
                    f"that could not be updated: {missing[:12]}"
                )
            extend_sheet_dimension(str(sheet_path), "K")
            hide_column_in_sheet(str(sheet_path), 11)  # column K

        # Only the private lookup is embedded. Diagnostics stay in the sync
        # log rather than adding an owner-visible worksheet.
        add_sheet(tempdir, "_PriceLookup", build_pricelookup_xml(rows), hidden=True)

        repack_zip(tempdir, CATALOGUE)
        tempdir = None
    finally:
        if tempdir:
            shutil.rmtree(tempdir, ignore_errors=True)

    print()
    print(f"  Match summary:")
    for k in ("persisted", "strong", "spec", "unique", "price-bootstrap",
              "equivalent-duplicate", "deterministic", "ambiguous",
              "unmatched", "no-data"):
        print(f"    {k:12s}: {counts[k]}")
    print(f"  Auto-linked (VLOOKUP):  {counts['auto_linked']}")
    print(f"  Unresolved rows:         {len(review_rows)}")
    if matcher.broken_links:
        print(f"  WARNING: {len(matcher.broken_links)} saved links pointed to removed master rows")
    for item in review_rows[:12]:
        print(
            f"    {item['confidence']}: {item['sheet']}!{item['row']} - "
            f"{item['metal']} / {item['spec']} / {item['size']}"
        )
    if len(review_rows) > 12:
        print(f"    ... plus {len(review_rows) - 12} more unresolved rows")
    print("  OK Excel-compatible catalogue written")


if __name__ == "__main__":
    wire_catalogue()
